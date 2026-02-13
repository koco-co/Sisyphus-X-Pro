"""Report export service for generating PDF, Excel, and HTML reports."""

from io import BytesIO

from fastapi import HTTPException, status
from sqlalchemy.ext.asyncio import AsyncSession

try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import inch
    from reportlab.platypus import (
        Paragraph,
        SimpleDocTemplate,
        Spacer,
        Table,
        TableStyle,
    )

    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False
    print("Warning: ReportLab not available. PDF export will be disabled.")

try:
    from openpyxl import Workbook
    from openpyxl.cell import MergedCell
    from openpyxl.styles import Alignment, Font, PatternFill

    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    MergedCell = None  # type: ignore
    print("Warning: openpyxl not available. Excel export will be disabled.")

try:
    from jinja2 import Template

    JINJA2_AVAILABLE = True
except ImportError:
    JINJA2_AVAILABLE = False
    print("Warning: Jinja2 not available. HTML export will be disabled.")

from app.services.report_service import ReportService


class ReportExportService:
    """Service for exporting test reports."""

    def __init__(self, session: AsyncSession) -> None:
        """Initialize export service.

        Args:
            session: Database session
        """
        self.session = session
        self.report_service = ReportService(session)

    async def export_pdf(self, report_id: int, include_details: bool = True) -> bytes:
        """Export report as PDF.

        Args:
            report_id: Report ID
            include_details: Include execution details

        Returns:
            PDF file bytes

        Raises:
            HTTPException: If ReportLab is not available
        """
        if not REPORTLAB_AVAILABLE:
            raise HTTPException(
                status_code=status.HTTP_501_NOT_IMPLEMENTED,
                detail="PDF export not available. Install reportlab.",
            )

        # Get report data
        details = await self.report_service.get_report_details(report_id)
        report = details["report"]
        scenarios = details["scenarios"]

        # Create PDF buffer
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter)
        styles = getSampleStyleSheet()
        story = []

        # Title
        title_style = ParagraphStyle(
            "CustomTitle",
            parent=styles["Heading1"],
            fontSize=24,
            textColor=colors.HexColor("#2c3e50"),
            spaceAfter=30,
        )
        story.append(Paragraph(f"Test Report #{report.id}", title_style))
        story.append(Spacer(1, 12))

        # Summary table
        summary_data = [
            ["Execution ID", report.execution_id],
            ["Status", report.status],
            ["Environment", report.environment_name],
            ["Started At", report.started_at.strftime("%Y-%m-%d %H:%M:%S")],
            ["Duration", f"{report.duration_seconds:.2f}s" if report.duration_seconds else "N/A"],
            ["Total Scenarios", str(report.total_scenarios)],
            ["Passed", f"{report.passed} ({self._pass_rate(report):.1f}%)"],
            ["Failed", str(report.failed)],
            ["Skipped", str(report.skipped)],
        ]

        summary_table = Table(summary_data, colWidths=[2 * inch, 3 * inch])
        summary_table.setStyle(
            TableStyle([
                ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#ecf0f1")),
                ("TEXTCOLOR", (0, 0), (0, -1), colors.HexColor("#2c3e50")),
                ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
                ("FONTSIZE", (0, 0), (-1, -1), 10),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
                ("GRID", (0, 0), (-1, -1), 1, colors.grey),
            ])
        )
        story.append(summary_table)
        story.append(Spacer(1, 24))

        # Scenarios details
        if include_details and scenarios:
            story.append(Paragraph("Scenario Details", styles["Heading2"]))
            story.append(Spacer(1, 12))

            for idx, scenario in enumerate(scenarios, 1):
                # Scenario header
                scenario_style = ParagraphStyle(
                    "Scenario",
                    parent=styles["Heading3"],
                    fontSize=14,
                    textColor=colors.HexColor("#34495e"),
                )
                story.append(
                    Paragraph(
                        f"Scenario {idx}: {scenario['scenario_id']}",
                        scenario_style,
                    )
                )
                story.append(Spacer(1, 6))

                # Scenario info
                scenario_data = [
                    ["Status", scenario["status"]],
                    ["Elapsed", f"{scenario['elapsed_ms']}ms" if scenario["elapsed_ms"] else "N/A"],
                ]
                if scenario["error_message"]:
                    scenario_data.append(["Error", scenario["error_message"]])

                scenario_table = Table(scenario_data, colWidths=[1.5 * inch, 4 * inch])
                scenario_table.setStyle(
                    TableStyle([
                        ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#ecf0f1")),
                        ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                        ("FONTSIZE", (0, 0), (-1, -1), 9),
                        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
                        ("GRID", (0, 0), (-1, -1), 1, colors.lightgrey),
                    ])
                )
                story.append(scenario_table)
                story.append(Spacer(1, 12))

                # Steps table
                if scenario["steps"]:
                    steps_data = [["#", "Step ID", "Status", "Elapsed (ms)", "Error"]]
                    for step in scenario["steps"]:
                        steps_data.append([
                            str(step["sort_order"]),
                            str(step["step_id"]),
                            step["status"],
                            str(step["elapsed_ms"]) if step["elapsed_ms"] else "N/A",
                            step["error_message"] or "",
                        ])

                    steps_table = Table(steps_data, colWidths=[0.5 * inch, 1 * inch, 1 * inch, 1 * inch, 2.5 * inch])
                    steps_table.setStyle(
                        TableStyle([
                            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#3498db")),
                            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                            ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                            ("FONTSIZE", (0, 0), (-1, -1), 8),
                            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
                            ("GRID", (0, 0), (-1, -1), 1, colors.grey),
                        ])
                    )
                    story.append(steps_table)
                    story.append(Spacer(1, 12))

                if idx < len(scenarios):
                    story.append(Spacer(1, 12))

        # Build PDF
        doc.build(story)
        buffer.seek(0)
        return buffer.read()

    async def export_excel(self, report_id: int, include_details: bool = True) -> bytes:
        """Export report as Excel.

        Args:
            report_id: Report ID
            include_details: Include execution details

        Returns:
            Excel file bytes

        Raises:
            HTTPException: If openpyxl is not available
        """
        if not OPENPYXL_AVAILABLE:
            raise HTTPException(
                status_code=status.HTTP_501_NOT_IMPLEMENTED,
                detail="Excel export not available. Install openpyxl.",
            )

        # Get report data
        details = await self.report_service.get_report_details(report_id)
        report = details["report"]
        scenarios = details["scenarios"]

        # Create workbook
        wb = Workbook()
        ws = wb.active
        if ws is None:
            ws = wb.create_sheet("Test Report")
        else:
            ws.title = "Test Report"

        # Title
        ws["A1"] = f"Test Report #{report.id}"
        ws["A1"].font = Font(size=16, bold=True, color="FFFFFF")
        ws["A1"].fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
        ws.merge_cells("A1:B1")

        # Summary
        row = 3
        summary_data = [
            ("Execution ID", report.execution_id),
            ("Status", report.status),
            ("Environment", report.environment_name),
            ("Started At", report.started_at.strftime("%Y-%m-%d %H:%M:%S")),
            ("Duration", f"{report.duration_seconds:.2f}s" if report.duration_seconds else "N/A"),
            ("Total Scenarios", str(report.total_scenarios)),
            ("Passed", str(report.passed)),
            ("Failed", str(report.failed)),
            ("Skipped", str(report.skipped)),
        ]

        for label, value in summary_data:
            ws[f"A{row}"] = label
            ws[f"A{row}"].font = Font(bold=True)
            ws[f"B{row}"] = value
            row += 1

        row += 1

        # Scenarios
        if include_details and scenarios:
            ws[f"A{row}"] = "Scenario Details"
            ws[f"A{row}"].font = Font(size=12, bold=True)
            row += 2

            for idx, scenario in enumerate(scenarios, 1):
                # Scenario header
                ws[f"A{row}"] = f"Scenario {idx}"
                ws[f"A{row}"].font = Font(bold=True, color="FFFFFF")
                ws[f"A{row}"].fill = PatternFill(start_color="3498DB", end_color="3498DB", fill_type="solid")
                ws.merge_cells(f"A{row}:B{row}")
                row += 1

                # Scenario info
                ws[f"A{row}"] = "Status"
                ws[f"B{row}"] = scenario["status"]
                row += 1
                ws[f"A{row}"] = "Elapsed (ms)"
                ws[f"B{row}"] = scenario.get("elapsed_ms") or "N/A"
                row += 1
                if scenario.get("error_message"):
                    ws[f"A{row}"] = "Error"
                    ws[f"B{row}"] = scenario["error_message"]
                    row += 1

                row += 1

                # Steps
                if scenario["steps"]:
                    ws[f"A{row}"] = "Steps"
                    ws[f"A{row}"].font = Font(bold=True)
                    row += 1

                    # Header row
                    headers = ["#", "Step ID", "Status", "Elapsed (ms)", "Error"]
                    for col, header in enumerate(headers, 1):
                        cell = ws.cell(row=row, column=col)
                        if MergedCell is None or not isinstance(cell, MergedCell):
                            cell.value = header  # type: ignore[arg-type]
                            cell.font = Font(bold=True, color="FFFFFF")
                            cell.fill = PatternFill(start_color="2ECC71", end_color="2ECC71", fill_type="solid")
                            cell.alignment = Alignment(horizontal="center")
                    row += 1

                    # Step rows
                    for step in scenario["steps"]:
                        for col_idx, value in enumerate([
                            step["sort_order"],
                            step["step_id"],
                            step["status"],
                            step.get("elapsed_ms") or "N/A",
                            step.get("error_message") or "",
                        ], 1):
                            cell = ws.cell(row=row, column=col_idx)
                            if MergedCell is None or not isinstance(cell, MergedCell):
                                cell.value = value  # type: ignore[arg-type]
                        row += 1

                row += 1

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            # Get column letter from first cell
            if column:
                first_cell = column[0]
                # Safely get column letter
                if hasattr(first_cell, 'column_letter') and not isinstance(first_cell, MergedCell if MergedCell else object):
                    column_letter = first_cell.column_letter  # type: ignore[attr-defined]
                else:
                    # Fallback: calculate column letter from index
                    col_idx = list(ws.columns).index(column) + 1
                    column_letter = chr(64 + col_idx) if col_idx <= 26 else 'A'

                for cell in column:
                    try:
                        if (MergedCell is None or not isinstance(cell, MergedCell)) and cell.value:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                    except Exception:
                        pass
                adjusted_width = (max_length + 2) * 1.2
                ws.column_dimensions[column_letter].width = min(adjusted_width, 50)

        # Save to buffer
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.read()

    async def export_html(self, report_id: int, include_details: bool = True) -> str:
        """Export report as HTML.

        Args:
            report_id: Report ID
            include_details: Include execution details

        Returns:
            HTML content

        Raises:
            HTTPException: If Jinja2 is not available
        """
        if not JINJA2_AVAILABLE:
            raise HTTPException(
                status_code=status.HTTP_501_NOT_IMPLEMENTED,
                detail="HTML export not available. Install jinja2.",
            )

        # Get report data
        details = await self.report_service.get_report_details(report_id)
        report = details["report"]
        scenarios = details["scenarios"]

        # Calculate pass rate
        pass_rate = self._pass_rate(report)

        # HTML template
        template_str = """
<!DOCTYPE html>
<html lang="zh-CN">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Test Report #{{ report.id }}</title>
    <style>
        body {
            font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif;
            line-height: 1.6;
            color: #333;
            max-width: 1200px;
            margin: 0 auto;
            padding: 20px;
            background: #f5f5f5;
        }
        .container {
            background: white;
            padding: 30px;
            border-radius: 8px;
            box-shadow: 0 2px 4px rgba(0,0,0,0.1);
        }
        h1 {
            color: #2c3e50;
            border-bottom: 2px solid #3498db;
            padding-bottom: 10px;
        }
        .summary {
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(200px, 1fr));
            gap: 20px;
            margin: 20px 0;
        }
        .summary-card {
            background: #ecf0f1;
            padding: 15px;
            border-radius: 6px;
            border-left: 4px solid #3498db;
        }
        .summary-card h3 {
            margin: 0 0 10px 0;
            color: #7f8c8d;
            font-size: 14px;
        }
        .summary-card .value {
            font-size: 24px;
            font-weight: bold;
            color: #2c3e50;
        }
        .status-badge {
            display: inline-block;
            padding: 4px 12px;
            border-radius: 12px;
            font-size: 12px;
            font-weight: bold;
        }
        .status-completed {
            background: #d4edda;
            color: #155724;
        }
        .status-failed {
            background: #f8d7da;
            color: #721c24;
        }
        .status-passed {
            background: #d4edda;
            color: #155724;
        }
        .status-failed-badge {
            background: #f8d7da;
            color: #721c24;
        }
        .status-skipped {
            background: #fff3cd;
            color: #856404;
        }
        table {
            width: 100%;
            border-collapse: collapse;
            margin: 20px 0;
        }
        th, td {
            padding: 12px;
            text-align: left;
            border-bottom: 1px solid #ddd;
        }
        th {
            background: #34495e;
            color: white;
            font-weight: bold;
        }
        tr:hover {
            background: #f8f9fa;
        }
        .scenario {
            margin: 30px 0;
            border: 1px solid #ddd;
            border-radius: 6px;
            overflow: hidden;
        }
        .scenario-header {
            background: #3498db;
            color: white;
            padding: 15px;
            font-weight: bold;
        }
        .scenario-body {
            padding: 15px;
        }
        .progress-bar {
            width: 100%;
            height: 30px;
            background: #ecf0f1;
            border-radius: 15px;
            overflow: hidden;
            display: flex;
            margin: 20px 0;
        }
        .progress-passed {
            background: #2ecc71;
            display: flex;
            align-items: center;
            justify-content: center;
            color: white;
            font-size: 12px;
        }
        .progress-failed {
            background: #e74c3c;
            display: flex;
            align-items: center;
            justify-content: center;
            color: white;
            font-size: 12px;
        }
        .progress-skipped {
            background: #f39c12;
            display: flex;
            align-items: center;
            justify-content: center;
            color: white;
            font-size: 12px;
        }
        .footer {
            margin-top: 30px;
            padding-top: 20px;
            border-top: 1px solid #ddd;
            color: #7f8c8d;
            font-size: 12px;
        }
    </style>
</head>
<body>
    <div class="container">
        <h1>Test Report #{{ report.id }}</h1>

        <div class="summary">
            <div class="summary-card">
                <h3>Status</h3>
                <div class="value">
                    <span class="status-badge status-{{ report.status }}">{{ report.status }}</span>
                </div>
            </div>
            <div class="summary-card">
                <h3>Total Scenarios</h3>
                <div class="value">{{ report.total_scenarios }}</div>
            </div>
            <div class="summary-card">
                <h3>Passed</h3>
                <div class="value" style="color: #2ecc71;">{{ report.passed }}</div>
            </div>
            <div class="summary-card">
                <h3>Failed</h3>
                <div class="value" style="color: #e74c3c;">{{ report.failed }}</div>
            </div>
            <div class="summary-card">
                <h3>Skipped</h3>
                <div class="value" style="color: #f39c12;">{{ report.skipped }}</div>
            </div>
            <div class="summary-card">
                <h3>Pass Rate</h3>
                <div class="value">{{ "%.1f"|format(pass_rate) }}%</div>
            </div>
        </div>

        {% if report.duration_seconds %}
        <div class="summary-card">
            <h3>Duration</h3>
            <div class="value">{{ "%.2f"|format(report.duration_seconds) }}s</div>
        </div>
        {% endif %}

        <div class="progress-bar">
            <div class="progress-passed" style="width: {{ (report.passed / report.total_scenarios * 100) if report.total_scenarios > 0 else 0 }}%">
                {{ report.passed }} passed
            </div>
            <div class="progress-failed" style="width: {{ (report.failed / report.total_scenarios * 100) if report.total_scenarios > 0 else 0 }}%">
                {{ report.failed }} failed
            </div>
            <div class="progress-skipped" style="width: {{ (report.skipped / report.total_scenarios * 100) if report.total_scenarios > 0 else 0 }}%">
                {{ report.skipped }} skipped
            </div>
        </div>

        <table>
            <tr>
                <th>Metric</th>
                <th>Value</th>
            </tr>
            <tr>
                <td>Execution ID</td>
                <td><code>{{ report.execution_id }}</code></td>
            </tr>
            <tr>
                <td>Environment</td>
                <td>{{ report.environment_name }}</td>
            </tr>
            <tr>
                <td>Started At</td>
                <td>{{ report.started_at.strftime('%Y-%m-%d %H:%M:%S') }}</td>
            </tr>
            {% if report.finished_at %}
            <tr>
                <td>Finished At</td>
                <td>{{ report.finished_at.strftime('%Y-%m-%d %H:%M:%S') }}</td>
            </tr>
            {% endif %}
        </table>

        {% if include_details and scenarios %}
        <h2>Scenario Details</h2>
        {% for scenario in scenarios %}
        <div class="scenario">
            <div class="scenario-header">
                Scenario #{{ loop.index }}: ID={{ scenario.scenario_id }}
            </div>
            <div class="scenario-body">
                <table>
                    <tr>
                        <td>Status</td>
                        <td><span class="status-badge status-{{ scenario.status }}">{{ scenario.status }}</span></td>
                    </tr>
                    <tr>
                        <td>Elapsed</td>
                        <td>{{ scenario.elapsed_ms }}ms</td>
                    </tr>
                    {% if scenario.error_message %}
                    <tr>
                        <td>Error</td>
                        <td style="color: #e74c3c;">{{ scenario.error_message }}</td>
                    </tr>
                    {% endif %}
                </table>

                {% if scenario.steps %}
                <h3>Steps</h3>
                <table>
                    <thead>
                        <tr>
                            <th>#</th>
                            <th>Step ID</th>
                            <th>Status</th>
                            <th>Elapsed (ms)</th>
                            <th>Error</th>
                        </tr>
                    </thead>
                    <tbody>
                        {% for step in scenario.steps %}
                        <tr>
                            <td>{{ step.sort_order }}</td>
                            <td>{{ step.step_id }}</td>
                            <td><span class="status-badge status-{{ step.status }}">{{ step.status }}</span></td>
                            <td>{{ step.elapsed_ms or 'N/A' }}</td>
                            <td>{{ step.error_message or '' }}</td>
                        </tr>
                        {% endfor %}
                    </tbody>
                </table>
                {% endif %}
            </div>
        </div>
        {% endfor %}
        {% endif %}

        <div class="footer">
            Generated by Sisyphus-X-Pro on {{ report.created_at.strftime('%Y-%m-%d %H:%M:%S') }}
        </div>
    </div>
</body>
</html>
        """

        template = Template(template_str)
        return template.render(report=report, scenarios=scenarios, pass_rate=pass_rate, include_details=include_details)

    def _pass_rate(self, report) -> float:
        """Calculate pass rate.

        Args:
            report: Test report

        Returns:
            Pass rate as percentage
        """
        if report.total_scenarios == 0:
            return 0.0
        return (report.passed / report.total_scenarios) * 100
